# Vista d'admin per exportar pronòstics a Excel
# Afegir a euro2024/joc/views/exporta_pronostics.py

from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from joc.models import Jugador, PronosticPartit, Partit
from django.conf import settings
from django.db.models import Q


@staff_member_required
def exporta_pronostics(request):
    wb = Workbook()
    wb.remove(wb.active)  # Treiem el full per defecte

    # Capçalera estil
    estil_cap = Font(bold=True, color='FFFFFF')
    fons_cap = PatternFill('solid', start_color='1F3864')
    centrat = Alignment(horizontal='center')

    # Obtenim tots els jugadors excepte l'admin
    jugadors = Jugador.objects.filter(
        usuari__is_active=True
    ).filter(
        ~Q(id=settings.ID_ADMIN)
    ).order_by('usuari__username')

    # Obtenim tots els partits ordenats
    partits = Partit.objects.all().order_by('id')

    for jugador in jugadors:
        # Nom del full = nom d'usuari (màx 31 caràcters, sense caràcters especials)
        nom_full = jugador.usuari.username[:31]
        nom_full = nom_full.replace('/', '-').replace('\\', '-').replace('*', '-')
        nom_full = nom_full.replace('?', '-').replace('[', '-').replace(']', '-')
        ws = wb.create_sheet(title=nom_full)

        # Capçaleres
        caps = ['Grup', 'Data', 'Equip 1', 'Gols 1', 'Gols 2', 'Equip 2', 'Empat']
        for col, cap in enumerate(caps, 1):
            cel = ws.cell(row=1, column=col, value=cap)
            cel.font = estil_cap
            cel.fill = fons_cap
            cel.alignment = centrat

        # Amplades de columnes
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 22
        ws.column_dimensions['G'].width = 8

        # Pronòstics del jugador
        pronostics = {
            pp.partit_id: pp
            for pp in PronosticPartit.objects.filter(jugador=jugador).select_related(
                'partit', 'equip1', 'equip2', 'partit__grup', 'partit__estadi'
            )
        }

        fila = 2
        grup_actual = None
        for partit in partits:
            pp = pronostics.get(partit.id)
            if pp is None:
                continue

            # Separador de grup
            if partit.grup.nom != grup_actual:
                grup_actual = partit.grup.nom

            equip1 = pp.equip1.nom if pp.equip1 else (partit.equip1.nom if partit.equip1 else '?')
            equip2 = pp.equip2.nom if pp.equip2 else (partit.equip2.nom if partit.equip2 else '?')
            gols1 = pp.gols1 if pp.gols1 >= 0 else '-'
            gols2 = pp.gols2 if pp.gols2 >= 0 else '-'
            empat = pp.empat if pp.empat else ''

            ws.cell(row=fila, column=1, value=partit.grup.nom)
            ws.cell(row=fila, column=2, value=str(partit.diaihora))
            ws.cell(row=fila, column=3, value=equip1)
            ws.cell(row=fila, column=4, value=gols1).alignment = centrat
            ws.cell(row=fila, column=5, value=gols2).alignment = centrat
            ws.cell(row=fila, column=6, value=equip2)
            ws.cell(row=fila, column=7, value=empat).alignment = centrat

            # Files alternades
            if fila % 2 == 0:
                for col in range(1, 8):
                    ws.cell(row=fila, column=col).fill = PatternFill('solid', start_color='DCE6F1')

            fila += 1

    # Resposta HTTP amb el fitxer Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="pronostics_mundial2026.xlsx"'
    wb.save(response)
    return response
